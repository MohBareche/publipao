import tkinter as tk
from tkinter import *
from tkinter import font, messagebox as mb, filedialog as fd, IntVar
from threading import Thread
import ttkbootstrap as ttk
from ttkbootstrap import Style

import docx2pdf
import os
import sys
import re
import glob
import shutil
import win32com.client

from openpyxl import load_workbook
from docxtpl import DocxTemplate
from pathlib import Path
from PyPDF2 import PdfMerger, PdfReader

global data_name
global ent_list
global soum_list
global adj_list

to_list = []
from_list = []


def confirm_quitter():
    answer = mb.askyesno(title='Confirmation',
                         message='Êtes-vous sûr de vouloir quitter?')
    if answer:
        window.destroy()


def confirm_pub_tout():
    mb.showinfo(title='Confirmation',
                message="Publipostage réalisé avec succès.")


def select_data_file():
    filetypes = (
        ('Fichier Excel', '*.xlsx'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title='Choisir une base de données',
        initialdir='./data',
        filetypes=filetypes
    )
    data_name = Path(filename)
    global wb
    wb = load_workbook(data_name)

    # liste chargés de projet
    ws_charg_proj = wb['Chargés de projet']
    list_charg_proj = []
    for cell in ws_charg_proj['B'][1:]:
        if cell.value != 'None':
            list_charg_proj.append(cell.value)
        lbl_message.grid(row=0, column=2)
        cmb_nom_charg_projet['values'] = list_charg_proj
        cmb_nom_charg_projet.configure(state='readonly')

    # liste gestionnaires
    ws_gest = wb['Gestionnaires']
    list_gestionnaires = []
    for cell in ws_gest['A'][1:]:
        if cell.value != 'None':
            list_gestionnaires.append(cell.value)
        cmb_nom_gestionnaire['values'] = list_gestionnaires
        cmb_nom_gestionnaire.current(0)
        cmb_nom_gestionnaire.configure(state='readonly')

    # liste secrétaires
    list_secret = []
    for cell in ws_gest['E'][1:]:
        if cell.value != 'None':
            list_secret.append(cell.value)
        cmb_secretaire['values'] = list_secret
        cmb_secretaire.current(0)
        cmb_secretaire.configure(state='readonly')

    select_remerc_file()
    select_octroi_file()

    lbl_message.configure(
        text='Base de données chargée avec succès...', width=50, relief='groove', bootstyle='inverse-success')


def load_data():
    excel_filename = 'data/Registre des entrepreneurs.xlsx'
    path_current = os.getcwd()
    path = f"{path_current}\{excel_filename}"

    wb = load_workbook(path)
    sheet = wb.active

    list_values = list(sheet.values)
    cols = list_values[0]

    tree = ttk.Treeview(window, columns=cols, show='headings')
    tree.pack(expand=True, fill='y')

    for col_name in cols:
        tree.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        tree.insert('', tk.END, values=value_tuple)

    tree.column('Ville', width=160, anchor=tk.CENTER)
    tree.column('Représentant', width=130)
    tree.column('Code Postal', width=100, anchor=tk.CENTER)
    tree.column('Civilité', width=80, anchor=tk.CENTER)
    tree.column('Fonction', width=150)


def moveTo(from_list, to_list):
    selected_items = [from_list.get(idx) for idx in from_list.curselection()]
    for item in selected_items:
        to_list.insert(tk.END, item)
        from_list.delete(from_list.curselection())


def move_adj(from_list, to_list):
    selected_items = [from_list.get(idx) for idx in from_list.curselection()]
    for item in selected_items:
        to_list.insert(tk.END, item)
        from_list.delete(from_list.curselection())
        btn_adj_1.configure(state='disabled')
        btn_adj_2.configure(state='normal')


def back_adj(from_list, to_list):
    selected_items = [from_list.get(idx) for idx in from_list.curselection()]
    for item in selected_items:
        to_list.insert(tk.END, item)
        from_list.delete(from_list.curselection())
        btn_adj_1.configure(state='normal')
        btn_adj_2.configure(state='disabled')


def soum_to_adj(e):
    if not adj_list.get(0, tk.END):
        btn_adj_1.configure(state='normal')
    else:
        btn_adj_1.configure(state='disabled')
        btn_adj_2.configure(state='normal')


def adj_to_soum(e):
    if not adj_list.get(0, tk.END):
        btn_adj_1.configure(state='disabled')
        btn_adj_2.configure(state='disabled')
    else:
        btn_adj_2.configure(state='normal')


def move_all(f_list, t_list):
    all_items = f_list.get(0, tk.END)
    f_list.delete(0, tk.END)
    for item in all_items:
        t_list.insert(tk.END, item)


def dbl_moveTo(e):
    ind_list = ent_list.curselection()
    if ind_list:
        ind = ind_list[0]
        val = ent_list.get(ind)
        ent_list.delete(ind)
        soum_list.insert(tk.END, val)


def dbl_moveBack(e):
    ind_list = soum_list.curselection()
    if ind_list:
        ind = ind_list[0]
        val = soum_list.get(ind)
        soum_list.delete(ind)
        ent_list.insert(tk.END, val)


def show_list_ent(e):
    ent_list.delete(0, tk.END)
    soum_list.delete(0, tk.END)
    nom_charg_proj = cmb_nom_charg_projet.get()
    data = wb['Chargés de projet']

    for row in data.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value == nom_charg_proj:
                global discipline
                specialite = data.cell(row=cell.row, column=4).value
                discipline = specialite

                if specialite == 'Voirie':
                    sheet_voirie = wb['Voirie']
                    m_row = sheet_voirie.max_row
                    for i in range(2, m_row + 1):
                        nom_ent = sheet_voirie.cell(row=i, column=1)
                        ent_list.insert(tk.END, nom_ent.value)

                if specialite == 'Bâtiment':
                    list_ent_bat = wb['Bâtiment']
                    m_row = list_ent_bat.max_row
                    for i in range(2, m_row + 1):
                        nom_ent = list_ent_bat.cell(row=i, column=1)
                        ent_list.insert(tk.END, nom_ent.value)

                if specialite == 'APA':
                    list_ent_apa = wb['Paysage']
                    m_row = list_ent_apa.max_row
                    for i in range(2, m_row + 1):
                        nom_ent = list_ent_apa.cell(row=i, column=1)
                        ent_list.insert(tk.END, nom_ent.value)


def select_remerc_file():
    global doc_remerc_name
    doc_remerc_name = 'Lettre_remerciement.docx'
    return doc_remerc_name


def select_octroi_file():
    global doc_octroi_name
    doc_octroi_name = 'Lettre_octroi.docx'
    return doc_octroi_name


def select_pv_ouverture_file():
    global doc_pv_ouvert_name
    filetypes = (
        ('Fichier PDF', '*.pdf'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title="Sélectionner le PV d'ouverture",
        initialdir='./pv',
        filetypes=filetypes
    )
    doc_pv_ouvert_name = Path(filename).name
    if doc_pv_ouvert_name:
        lbl_pv_ouvert.configure(
            text="Procès verbal d'ouverture (OK)", bootstyle='SUCCESS')
        return doc_pv_ouvert_name


def select_pv_ca_file():
    global doc_pv_ca_name
    filetypes = (
        ('Fichier Word', '*.doc'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title='Sélectionner le PV du CA',
        initialdir='./pv',
        filetypes=filetypes
    )
    doc_pv_ca_name = Path(filename).name
    if doc_pv_ca_name:
        lbl_pv_ca.configure(text="Procès verbal CA (OK)", bootstyle='SUCCESS')
        return doc_pv_ca_name


def select_redac():
    global nom_redac
    if var_redac.get() == 0:
        cmb_secretaire.configure(state='readonly')
        nom_redac = cmb_secretaire.get()

    if var_redac.get() == 1:
        cmb_secretaire.configure(state='disabled')
        cmb_secretaire.config(foreground='silver')
        nom_redac = cmb_nom_charg_projet.get()
    return nom_redac


def get_secret_name(e):
    nom_redac = cmb_secretaire.get()
    return nom_redac


def initiales_gest(nom):
    cap = nom.split(' ')
    init = cap[0][0] + cap[1][0]
    return init


def initiales_redac(nom):
    cap = nom.split(' ')
    init = cap[0][0] + cap[1][0]
    return init.lower()


def erreur_msg():
    mb.showerror(title='Erreur',
                 message="Veuillez entrer les données manquantes.")


def gener_remerc():
    path = f'./gabarits/{doc_remerc_name}'
    doc = DocxTemplate(path)
    compagnies = {}
    ws = wb[discipline]
    for row in ws.iter_rows(min_row=2, values_only=True):
        company_name = row[0]
        company_data = {
            "nom_de_compagnie": row[0],
            "adresse": row[1],
            "ville": row[2],
            "code_postal": row[3],
            "courriel": row[4],
            "representant": row[5],
            "civilite": row[6],
            "fonction": row[7]
        }
        compagnies[company_name] = company_data

    ws_gestionnaires = wb['Gestionnaires']
    date = entry_cal.entry.get()
    titre_projet = entry_titre_projet.get()
    num_contrat = entry_num_contrat.get()
    nom_gest = cmb_nom_gestionnaire.get()
    init_gest = initiales_gest(nom_gest)

    if var_redac.get() == 0:
        nom_redac = cmb_secretaire.get()
    if var_redac.get() == 1:
        nom_redac = cmb_nom_charg_projet.get()

    init_redac = initiales_redac(nom_redac)

    for row in ws_gestionnaires.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == nom_gest:
                titre_gest = ws_gestionnaires.cell(
                    row=cell.row, column=2).value
                fonction_gest = ws_gestionnaires.cell(
                    row=cell.row, column=3).value

    path_folder = './output/remerciement'
    isExist = os.path.exists(path_folder)
    if isExist:
        remerc_folder = shutil.rmtree('./output/remerciement')
    pathDOC = './output/remerciement/DOC'

    os.makedirs(pathDOC)

    for ent in list(soum_list.get(0, tk.END)):
        doc.render({
            "date": date,
            "titre": titre_projet,
            "num_contrat": num_contrat,
            "nom_gestionnaire": nom_gest,
            "titre_gest": titre_gest,
            "fonction_gest": fonction_gest,
            "init_gest": init_gest,
            "init_redac": init_redac,
            "civilite": compagnies[ent]['civilite'],
            "representant": compagnies[ent]['representant'],
            "nom_de_compagnie": compagnies[ent]['nom_de_compagnie'],
            "adresse": compagnies[ent]['adresse'],
            "ville": compagnies[ent]['ville'],
            "code_postal": compagnies[ent]['code_postal'],
            "courriel": compagnies[ent]['courriel']
        })
        nom_comp = f'{compagnies[ent]["nom_de_compagnie"]}'
        nom_fichier = f"{num_contrat}_Lettre de remerciement - {nom_comp}.docx"

        doc.save(f'{pathDOC}/{nom_fichier}')

    docx2pdf.convert(pathDOC, '.')

    pv = f"./pv/{doc_pv_ouvert_name}"
    pdf_pv = open(pv, 'rb')

    pdfs = glob.glob('*.pdf')

    for pdf in pdfs:
        merger = PdfMerger()
        merger.append(pdf)
        merger.append(pdf_pv)
        name = pdf.split(".")[0]
        merger.write(f"{name}_fin.pdf")
        merger.close()

    pdf_folder = 'PDF'
    remerc_folder = './output/remerciement'
    pathPDF = os.path.join(remerc_folder, pdf_folder)
    os.makedirs(pathPDF)

    for f in glob.glob('./*_fin.pdf'):
        shutil.move(f, pathPDF)

    for f in os.listdir('./'):
        if f.endswith('.pdf'):
            os.remove(f)


def gener_octroi():
    path_gabarit = f'./gabarits/{doc_octroi_name}'
    doc = DocxTemplate(path_gabarit)
    compagnies = {}
    ws = wb[discipline]
    for row in ws.iter_rows(min_row=2, values_only=True):
        company_name = row[0]
        company_data = {
            "nom_de_compagnie": row[0],
            "adresse": row[1],
            "ville": row[2],
            "code_postal": row[3],
            "courriel": row[4],
            "representant": row[5],
            "civilite": row[6],
            "fonction": row[7]
        }
        compagnies[company_name] = company_data

        ws_gestionnaires = wb['Gestionnaires']

    pv_ca = f"./pv/{doc_pv_ca_name}"
    shutil.move(pv_ca, './')

    filename = doc_pv_ca_name
    filenamePDF = filename.split('.')[0]
    path = os.getcwd()
    in_file = f"{path}\{filename}"
    out_file = f"{path}\{filenamePDF}"

    wdFormatPDF = 17
    word = win32com.client.Dispatch('Word.Application')
    doc_doc = word.Documents.Open(in_file)
    doc_doc.SaveAs(out_file, FileFormat=wdFormatPDF)
    doc_doc.Close()
    word.Quit()
    shutil.move(in_file, './pv')

    reader = PdfReader(f"{out_file}.pdf")
    texte = reader.pages[0].extract_text()
    resolution = re.search(r"CA[\d]{2}\s[\d]{2}\s[\d]{2,4}", texte).group()
    date_resolution = re.search(
        r"[\d]{1,2}\s(?:janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\s[\d]{4}", texte).group()

    date = entry_cal.entry.get()
    titre_projet = entry_titre_projet.get()
    num_contrat = entry_num_contrat.get()
    num_ao = entry_num_ao.get()
    nom_gest = cmb_nom_gestionnaire.get()
    charg_projet = cmb_nom_charg_projet.get()
    init_gest = initiales_gest(nom_gest)

    if var_redac.get() == 0:
        nom_redac = cmb_secretaire.get()
    if var_redac.get() == 1:
        nom_redac = cmb_nom_charg_projet.get()

    init_redac = initiales_redac(nom_redac)

    ws_charg_proj = wb['Chargés de projet']
    for row in ws_charg_proj.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value == charg_projet:
                civ_charge_proj = ws_charg_proj.cell(
                    row=cell.row, column=1).value
                nom_charge_projet = ws_charg_proj.cell(
                    row=cell.row, column=2).value
                tel_charge_projet = ws_charg_proj.cell(
                    row=cell.row, column=3).value

    ws_gestionnaires = wb['Gestionnaires']
    for row in ws_gestionnaires.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == nom_gest:
                titre_gest = ws_gestionnaires.cell(
                    row=cell.row, column=2).value
                fonction_gest = ws_gestionnaires.cell(
                    row=cell.row, column=3).value

    path_folder = './output/octroi'
    isExist = os.path.exists(path_folder)
    if isExist:
        octroi_folder = shutil.rmtree('./output/octroi')
    pathDOC = './output/octroi/DOC'
    os.makedirs(pathDOC)

    for ent in list(adj_list.get(0, tk.END)):
        doc.render({
            "date": date,
            "titre": titre_projet,
            "num_contrat": num_contrat,
            "num_ao": num_ao,
            "nom_gestionnaire": nom_gest,
            "titre_gest": titre_gest,
            "fonction_gest": fonction_gest,
            "init_gest": init_gest,
            "init_redac": init_redac,
            "resolution": resolution,
            "date_resolution": date_resolution,
            "civ_charge_projet": civ_charge_proj,
            "nom_charge_projet": nom_charge_projet,
            "tel_charge_projet": tel_charge_projet,
            "civilite": compagnies[ent]['civilite'],
            "representant": compagnies[ent]['representant'],
            "nom_de_compagnie": compagnies[ent]['nom_de_compagnie'],
            "adresse": compagnies[ent]['adresse'],
            "ville": compagnies[ent]['ville'],
            "code_postal": compagnies[ent]['code_postal'],
            "courriel": compagnies[ent]['courriel']
        })
        global nom_comp_adj
        nom_comp_adj = f'{compagnies[ent]["nom_de_compagnie"]}'
        nom_fichier_doc = f"{num_contrat}_Lettre d'adjudication - {nom_comp_adj}.docx"
        doc.save(f'{pathDOC}/{nom_fichier_doc}')

    docx2pdf.convert(pathDOC, '.')

    pdfs = glob.glob('*.pdf')

    pdfs = [f for f in os.listdir() if f.endswith(".pdf")]

    merger = PdfMerger()

    for pdf in pdfs:
        merger.append(open(pdf, 'rb'))

    pdf_folder = 'PDF'
    octroi_folder = './output/octroi'
    pathPDF = os.path.join(octroi_folder, pdf_folder)
    os.makedirs(pathPDF)

    nom_fichier_pdf = f"{pathPDF}/{num_contrat}_Lettre d'adjudication - {nom_comp_adj}.pdf"
    with open(nom_fichier_pdf, 'wb') as fout:
        merger.write(fout)
        merger.close()

    for f in os.listdir('./'):
        if f.endswith('.pdf'):
            os.remove(f)


def gener_tout():
    gener_remerc()
    gener_octroi()


def enable_btn_folder():
    btn_open_folder.configure(state='normal')


def update_theme(e):
    window.style.theme_use(nom_theme.get())


def open_folder():
    path = './output'
    os.system(f'start {os.path.realpath(path)}')


def show_and_run(func):
    func()


def run_function(func, btn):
    progressbar = ttk.Progressbar(frame_progress, orient='horizontal',
                                  mode='indeterminate', length=200, bootstyle="info-striped")
    progressbar.grid(row=0, column=0, padx=10, pady=10)
    progressbar.start(interval=10)
    frame_progress.configure(text='Veuillez patienter . . .')
    show_and_run(func)
    enable_btn_folder()
    progressbar.destroy()
    frame_progress.configure(text='Statut')
    confirm_pub_tout()


def generer(func, btn):
    global titre_projet
    global num_contrat
    global num_ao
    
    titre_projet = entry_titre_projet.get()
    num_contrat = entry_num_contrat.get()
    num_ao = entry_num_ao.get()
    nom_charg_proj = cmb_nom_charg_projet.get()
    
    if (titre_projet and num_contrat and num_ao and nom_charg_proj):
        Thread(target=run_function, args=(func, btn)).start()
    else:
        erreur_msg()


def restart_program():
    python = sys.executable
    os.execl(python, python, * sys.argv)


def reinit():
    pass


window = ttk.Window(themename='darkly')
window_width = 1275
window_height = 725

screen_with = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
pos_x = (screen_with - window_width) // 2
pos_y = (screen_height - window_height) // 2
window.geometry(f"{window_width}x{window_height}+{pos_x}+{pos_y}")
window.title("Publipostage - Lettres de remerciement et lettre d'octroi")
window.resizable(0, 0)

# default font
window.defaultFont = font.nametofont("TkDefaultFont")
window.defaultFont.configure(family="Arial", size=11)

themes = window.style.theme_names()
nom_theme = ttk.StringVar(value=window.style.theme_use())

# ********************************************************************************************************************

frame_first = ttk.Frame(window)
frame_first.grid(row=0, column=0, columnspan=2, sticky='NEWS', padx=20, pady=7)

frame_data = ttk.LabelFrame(
    frame_first, text='Choix de la base de données', width=500, height=100)
frame_data.grid(row=0, column=0, sticky='W', padx=20, pady=7)

lbl_load_data = ttk.Label(frame_data, text='Choisir un fichier')
lbl_load_data.grid(row=0, column=0)
btn_load_data = ttk.Button(
    frame_data, text='Sélectionner...', command=select_data_file, bootstyle='PRIMARY, OUTLINE')
btn_load_data.grid(row=0, column=1)

lbl_message = ttk.Label(frame_data,
                        text="Choisir d'abord la base de données pour commencer.", width=50, relief='groove', bootstyle="inverse-danger")
lbl_message.grid(row=0, column=2)

# *************************************************************************************************
frame_progress = ttk.LabelFrame(frame_first, width=200, text='Statut')
frame_progress.grid(row=0, column=1, sticky='NEWS', padx=20, pady=7)

# *************************************************************************************************
frame_theme = ttk.LabelFrame(
    frame_first, text='Sélectionner un thème', width=300, height=100)
frame_theme.grid(row=0, column=2,  sticky='E', padx=20, pady=7)

for theme in themes:
    cmb_theme = ttk.Combobox(
        frame_theme, state='readonly', textvariable=nom_theme, values=themes)
    cmb_theme.current(11)
    cmb_theme.grid(row=0, column=0, padx=5, pady=10)

for widget in frame_data.winfo_children():
    widget.grid_configure(padx=10, pady=7)

for widget in frame_progress.winfo_children():
    widget.grid_configure(padx=10, pady=7)

for widget in frame_theme.winfo_children():
    widget.grid_configure(padx=10, pady=7)

# ********************************************************************************************************************

frame_info_projet = ttk.LabelFrame(
    window,  text='Informations sur le projet', width=700, height=100)
frame_info_projet.grid(row=1, column=0, columnspan=2,
                       sticky='NEWS', padx=20, pady=7)

lbl_titre_projet = ttk.Label(frame_info_projet, text='Titre du projet')
lbl_titre_projet.grid(row=0, column=0)
entry_titre_projet = ttk.Entry(frame_info_projet, width=70)
entry_titre_projet.grid(row=0, column=1)

lbl_num_contrat = ttk.Label(frame_info_projet, text='Numéro de contrat')
lbl_num_contrat.grid(row=0, column=2)
entry_num_contrat = ttk.Entry(frame_info_projet)
entry_num_contrat.grid(row=0, column=3)

lbl_num_ao = ttk.Label(frame_info_projet, text="Numéro d'appel d'offres")
lbl_num_ao.grid(row=0, column=4)
entry_num_ao = ttk.Entry(frame_info_projet)
entry_num_ao.grid(row=0, column=5)

for widget in frame_info_projet.winfo_children():
    widget.grid_configure(padx=10, pady=10)
# ********************************************************************************************************************
frame_info_charg_proj_sign_date = ttk.LabelFrame(
    window, text='Informations diverses [ Chargé de projet, signataire, rédacteur et date de rédaction ]', width=700, height=100)
frame_info_charg_proj_sign_date.grid(
    row=2, column=0, columnspan=2, sticky='NEWS', padx=20, pady=7)

lbl_nom_charg_projet = ttk.Label(
    frame_info_charg_proj_sign_date, text='Chargé(e) de projet')
lbl_nom_charg_projet.grid(row=0, column=0, sticky='s')
cmb_nom_charg_projet = ttk.Combobox(
    frame_info_charg_proj_sign_date, width=25, bootstyle='PRIMARY')
cmb_nom_charg_projet.grid(row=1, column=0, sticky='n')
cmb_nom_charg_projet.bind("<<ComboboxSelected>>", show_list_ent)

# informations sur le signataire (gestionnaire)
lbl_nom_gestionnaire = ttk.Label(
    frame_info_charg_proj_sign_date, text='Signataire (Gestionnaire)')
lbl_nom_gestionnaire.grid(row=0, column=1, sticky='s')
cmb_nom_gestionnaire = ttk.Combobox(frame_info_charg_proj_sign_date, width=25)
cmb_nom_gestionnaire.grid(row=1, column=1, sticky='n')

# date de rédaction
lbl_date = ttk.Label(frame_info_charg_proj_sign_date,
                     text="Date de rédaction")
lbl_date.grid(row=0, column=2, sticky='s')

entry_cal = ttk.DateEntry(frame_info_charg_proj_sign_date)
entry_cal.grid(row=1, column=2, sticky='n')

# rédacteur
frm_redacteur = ttk.LabelFrame(
    frame_info_charg_proj_sign_date, text='Rédacteur')
frm_redacteur.grid(row=0, column=3, rowspan=2, padx=10, pady=10, sticky='e')

var_redac = IntVar(None, 0)
rbtn_red = ttk.Radiobutton(frm_redacteur, text='Secrétaire', bootstyle="INFO",
                           variable=var_redac, value=0, command=select_redac)
rbtn_red.grid(row=0, column=1, sticky='w', padx=10, pady=10)

rbtn_red = ttk.Radiobutton(frm_redacteur, text='Chargé(e) de projet', bootstyle="INFO",
                           variable=var_redac, value=1, command=select_redac)
rbtn_red.grid(row=1, column=1, sticky='w', padx=10, pady=10)

cmb_secretaire = ttk.Combobox(frm_redacteur, width=25, bootstyle='DARK')
cmb_secretaire.grid(row=0, column=2, sticky='e', padx=10)

for widget in frame_info_charg_proj_sign_date.winfo_children():
    widget.grid_configure(padx=15, pady=10)

# ********************************************************************************************************************
frame_soumission = ttk.LabelFrame(
    window, text='Informations sur les soumissionnaires', width=700, height=200)
frame_soumission.grid(row=3, column=0, columnspan=2,
                      sticky='NEWS', padx=20, pady=7)

lbl_list_ent = ttk.Label(frame_soumission, text='Liste des entrepreneurs')
lbl_list_ent.grid(row=0, column=0)
ent_list = tk.Listbox(frame_soumission,
                      width=40, font=('Arial', 10))
ent_list.grid(row=1, column=0)

frame_group_btn1 = ttk.Frame(frame_soumission)
frame_group_btn1.grid(row=1, column=1, rowspan=4)

btn_1 = ttk.Button(frame_group_btn1, text='>', bootstyle='DANGER, OUTLINE',
                   width=3, command=lambda: moveTo(ent_list, soum_list))
btn_1.grid(row=0, column=0, pady=5)

btn_2 = ttk.Button(frame_group_btn1, text='>>', bootstyle='DANGER, OUTLINE',
                   width=3, command=lambda: move_all(ent_list, soum_list))
btn_2.grid(row=1, column=0, pady=5)

btn_3 = ttk.Button(frame_group_btn1, text='<', bootstyle='DANGER, OUTLINE',
                   width=3, command=lambda: moveTo(soum_list, ent_list))
btn_3.grid(row=2, column=0, pady=5)

btn_4 = ttk.Button(frame_group_btn1, text='<<', bootstyle='DANGER, OUTLINE',
                   width=3, command=lambda: move_all(soum_list, ent_list))
btn_4.grid(row=3, column=0, pady=5)

lbl_list_soum = ttk.Label(frame_soumission, text='Liste des soumissionnaires')
lbl_list_soum.grid(row=0, column=2)
soum_list = tk.Listbox(frame_soumission, width=40, font=('Arial', 10))
soum_list.grid(row=1, column=2)

frame_group_btn2 = ttk.Frame(frame_soumission)
frame_group_btn2.grid(row=1, column=3)

btn_adj_1 = ttk.Button(frame_group_btn2, text='Octroyer>', state='disabled', bootstyle='SUCCESS, OUTLINE',
                       width=10, command=lambda: [move_adj(soum_list, adj_list), soum_to_adj])
btn_adj_1.grid(row=0, column=0, pady=5)

btn_adj_2 = ttk.Button(frame_group_btn2, text='<Retirer', state='disabled', bootstyle='SUCCESS, OUTLINE',
                       width=10, command=lambda: [back_adj(adj_list, soum_list), adj_to_soum])
btn_adj_2.grid(row=1, column=0, pady=5)

lbl_adj = ttk.Label(frame_soumission, text='Entreprise adjugée')
lbl_adj.grid(row=0, column=4)
adj_list = tk.Listbox(frame_soumission, width=40, font=('Arial', 10))

for widget in frame_soumission.winfo_children():
    widget.grid_configure(padx=10, pady=0)

adj_list.grid(row=1, column=4, pady=10)
# ********************************************************************************************************************
frame_remerc_octroi = ttk.LabelFrame(
    window, text='Lettres de remerciement et octroi', width=300, height=200)
frame_remerc_octroi.grid(row=4, column=0, sticky='NEWS', padx=20, pady=7)

frame_remerc = ttk.Frame(
    frame_remerc_octroi)
frame_remerc.grid(row=0, column=0, sticky='N', padx=20, pady=5)

# PV Ouverture Remerciements
lbl_pv_ouvert = ttk.Label(
    frame_remerc, text="Procès verbal d'ouverture (.pdf)", width=30)
lbl_pv_ouvert.grid(row=0, column=0, stick='W')

btn_pv_ouvert = ttk.Button(
    frame_remerc, text='Sélectionner...', bootstyle='INFO, OUTLINE', command=select_pv_ouverture_file, width=15)
btn_pv_ouvert.grid(row=0, column=1, sticky='W')

btn_gen_remerc = ttk.Button(
    frame_remerc, text='Générer les lettres de remerciement', bootstyle='SUCCESS', width=35,
    command=lambda: generer(gener_remerc, btn_gen_remerc))
btn_gen_remerc.grid(row=0, column=2, sticky='E')

for widget in frame_remerc.winfo_children():
    widget.grid_configure(padx=10, pady=10)

frame_octroi = ttk.Frame(
    frame_remerc_octroi)
frame_octroi.grid(row=1, column=0, sticky='S', padx=20, pady=5)
# PV CA Octroi
lbl_pv_ca = ttk.Label(frame_octroi, text='Procès verbal CA (.doc)', width=30)
lbl_pv_ca.grid(row=0, column=0, sticky='W')
btn_pv_ca = ttk.Button(
    frame_octroi, text='Sélectionner...', bootstyle='INFO, OUTLINE', command=select_pv_ca_file, width=15)
btn_pv_ca.grid(row=0, column=1, sticky='W')

btn_gen_octroi = ttk.Button(
    frame_octroi, text="Générer la lettre d'octroi", bootstyle='SUCCESS', width=35,
    command=lambda: generer(gener_octroi, btn_gen_octroi))
btn_gen_octroi.grid(row=0, column=2, sticky="E")

for widget in frame_octroi.winfo_children():
    widget.grid_configure(padx=10, pady=10)
# ********************************************************************************************************************
frame_btns = ttk.Frame(window, width=600, height=100)
frame_btns.grid(row=4, column=1, padx=20, pady=7, sticky='E')

btn_generer_tout = ttk.Button(
    frame_btns, text='Générer tout (remerciement et octroi)', width=45, bootstyle='PRIMARY',
    command=lambda: generer(gener_tout, btn_generer_tout))
btn_generer_tout.grid(row=0, column=0, columnspan=2)

frm_btns = ttk.Frame(frame_btns, width=600)
frm_btns.grid(row=1, column=0, padx=10, pady=5, sticky='NEWS')

btn_reinit = ttk.Button(frm_btns, text='Réinitialiser',
                        bootstyle='WARNING', width=25, command=reinit)
btn_reinit.grid(row=1, column=0, padx=5, sticky='W')

btn_open_folder = ttk.Button(frm_btns, text='Explorer', state='disabled',
                             bootstyle='INFO', width=15, command=open_folder)
btn_open_folder.grid(row=1, column=1, sticky='E')

btn_quitter = ttk.Button(frame_btns, text='Quitter',
                         width=45, bootstyle='DANGER', command=confirm_quitter)
btn_quitter.grid(row=2, column=0, columnspan=2)

for widget in frame_btns.winfo_children():
    widget.grid_configure(padx=10, pady=7)


ent_list.bind('<Double-Button>', dbl_moveTo)
soum_list.bind('<Double-Button>', dbl_moveBack)
soum_list.bind('<<ListboxSelect>>', soum_to_adj)
adj_list.bind('<<ListboxSelect>>', adj_to_soum)
cmb_secretaire.bind('<<ComboboxSelected>>', get_secret_name)
cmb_theme.bind('<<ComboboxSelected>>', update_theme)

window.mainloop()
